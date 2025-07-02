"""
Servicios para generar reportes en diferentes formatos
"""
import os
import tempfile
from typing import Dict, List, Any, Optional, Union
from datetime import datetime, date
from decimal import Decimal
import pandas as pd
import json
from django.conf import settings
from django.utils import timezone
from django.template.loader import render_to_string
from django.http import HttpResponse, FileResponse
from django.core.files.storage import default_storage
from django.db.models import QuerySet
import logging

logger = logging.getLogger(__name__)


class ReportGeneratorService:
    """Servicio principal para generar reportes"""
    
    def __init__(self):
        self.temp_dir = tempfile.mkdtemp()
        self.generators = {
            'pdf': PDFReportGenerator(),
            'excel': ExcelReportGenerator(),
            'csv': CSVReportGenerator(),
            'json': JSONReportGenerator(),
        }
    
    def generate_report(
        self,
        template,
        data: Dict[str, Any],
        format_type: str = 'pdf',
        output_path: Optional[str] = None
    ) -> Dict[str, Any]:
        """
        Genera un reporte en el formato especificado
        
        Args:
            template: Plantilla de reporte
            data: Datos para el reporte
            format_type: Formato de salida ('pdf', 'excel', 'csv', 'json')
            output_path: Ruta de salida opcional
        
        Returns:
            Dict con información del archivo generado
        """
        try:
            generator = self.generators.get(format_type)
            if not generator:
                raise ValueError(f"Formato no soportado: {format_type}")
            
            # Generar nombre de archivo si no se especifica
            if not output_path:
                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                filename = f"{template.name}_{timestamp}.{format_type}"
                output_path = os.path.join(self.temp_dir, filename)
            
            # Generar el reporte
            result = generator.generate(template, data, output_path)
            
            return {
                'success': True,
                'file_path': result['file_path'],
                'file_size': result['file_size'],
                'mime_type': result['mime_type'],
                'filename': os.path.basename(result['file_path']),
                'generation_time': result.get('generation_time', 0)
            }
            
        except Exception as e:
            logger.error(f"Error generando reporte: {str(e)}")
            return {
                'success': False,
                'error': str(e)
            }
    
    def get_report_data(self, template, filters: Dict[str, Any]) -> Dict[str, Any]:
        """
        Obtiene los datos para el reporte basado en la plantilla y filtros
        """
        try:
            data_service = ReportDataService()
            return data_service.get_data_for_template(template, filters)
        except Exception as e:
            logger.error(f"Error obteniendo datos para reporte: {str(e)}")
            raise


class BaseReportGenerator:
    """Clase base para generadores de reportes"""
    
    def generate(self, template, data: Dict[str, Any], output_path: str) -> Dict[str, Any]:
        """Método que deben implementar los generadores específicos"""
        raise NotImplementedError
    
    def _get_file_info(self, file_path: str) -> Dict[str, Any]:
        """Obtiene información básica del archivo generado"""
        file_size = os.path.getsize(file_path) / (1024 * 1024)  # MB
        return {
            'file_path': file_path,
            'file_size': round(file_size, 2),
        }


class PDFReportGenerator(BaseReportGenerator):
    """Generador de reportes en PDF"""
    
    def generate(self, template, data: Dict[str, Any], output_path: str) -> Dict[str, Any]:
        """Genera reporte PDF usando ReportLab"""
        start_time = timezone.now()
        
        try:
            from reportlab.lib.pagesizes import letter, A4
            from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, Image
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.lib.units import inch
            from reportlab.lib import colors
            from reportlab.graphics.shapes import Drawing
            from reportlab.graphics.charts.linecharts import HorizontalLineChart
            from reportlab.graphics.charts.barcharts import VerticalBarChart
            
            # Crear documento PDF
            doc = SimpleDocTemplate(
                output_path,
                pagesize=A4,
                rightMargin=72,
                leftMargin=72,
                topMargin=72,
                bottomMargin=18
            )
            
            # Estilos
            styles = getSampleStyleSheet()
            title_style = ParagraphStyle(
                'CustomTitle',
                parent=styles['Heading1'],
                fontSize=16,
                spaceAfter=30,
                alignment=1  # Center
            )
            
            # Contenido del documento
            story = []
            
            # Título
            title = Paragraph(f"Reporte: {template.name}", title_style)
            story.append(title)
            story.append(Spacer(1, 12))
            
            # Información del reporte
            info_data = [
                ['Fecha de generación:', timezone.now().strftime('%d/%m/%Y %H:%M')],
                ['Período:', f"{data.get('date_from', '')} - {data.get('date_to', '')}"],
                ['Tipo de reporte:', template.get_report_type_display()],
            ]
            
            info_table = Table(info_data, colWidths=[2*inch, 3*inch])
            info_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (0, -1), colors.lightgrey),
                ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
                ('FONTSIZE', (0, 0), (-1, -1), 10),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 12),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ]))
            
            story.append(info_table)
            story.append(Spacer(1, 20))
            
            # Datos principales
            if 'main_data' in data and data['main_data']:
                self._add_data_table(story, data['main_data'], styles)
            
            # KPIs
            if 'kpis' in data and data['kpis']:
                self._add_kpis_section(story, data['kpis'], styles)
            
            # Gráficos
            if 'charts' in data and data['charts']:
                self._add_charts_section(story, data['charts'], styles)
            
            # Construir PDF
            doc.build(story)
            
            generation_time = (timezone.now() - start_time).total_seconds()
            
            result = self._get_file_info(output_path)
            result.update({
                'mime_type': 'application/pdf',
                'generation_time': generation_time
            })
            
            return result
            
        except Exception as e:
            logger.error(f"Error generando PDF: {str(e)}")
            raise
    
    def _add_data_table(self, story, data, styles):
        """Agrega una tabla de datos al reporte"""
        if not data:
            return
        
        # Título de la sección
        story.append(Paragraph("Datos del Reporte", styles['Heading2']))
        story.append(Spacer(1, 12))
        
        # Convertir datos a tabla
        if isinstance(data, list) and data:
            headers = list(data[0].keys())
            table_data = [headers]
            
            for row in data[:50]:  # Limitar a 50 filas para PDF
                table_data.append([str(row.get(col, '')) for col in headers])
            
            # Crear tabla
            table = Table(table_data)
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 10),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ]))
            
            story.append(table)
            story.append(Spacer(1, 20))
    
    def _add_kpis_section(self, story, kpis, styles):
        """Agrega sección de KPIs al reporte"""
        story.append(Paragraph("Indicadores Clave (KPIs)", styles['Heading2']))
        story.append(Spacer(1, 12))
        
        kpi_data = [['KPI', 'Valor', 'Estado']]
        
        for kpi in kpis:
            status_text = {
                'green': 'Bueno',
                'yellow': 'Advertencia', 
                'red': 'Crítico',
                'gray': 'Sin datos'
            }.get(kpi.get('status', 'gray'), 'Desconocido')
            
            kpi_data.append([
                kpi.get('name', ''),
                str(kpi.get('value', '')),
                status_text
            ])
        
        kpi_table = Table(kpi_data, colWidths=[2*inch, 1*inch, 1*inch])
        kpi_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.lightblue),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.black),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        
        story.append(kpi_table)
        story.append(Spacer(1, 20))
    
    def _add_charts_section(self, story, charts, styles):
        """Agrega sección de gráficos al reporte"""
        story.append(Paragraph("Gráficos y Análisis", styles['Heading2']))
        story.append(Spacer(1, 12))
        
        # Por ahora solo agregamos un placeholder
        # En una implementación completa integraríamos con matplotlib/plotly
        chart_text = Paragraph(
            "Los gráficos se incluirían aquí en una implementación completa.",
            styles['Normal']
        )
        story.append(chart_text)
        story.append(Spacer(1, 20))


class ExcelReportGenerator(BaseReportGenerator):
    """Generador de reportes en Excel"""
    
    def generate(self, template, data: Dict[str, Any], output_path: str) -> Dict[str, Any]:
        """Genera reporte Excel usando openpyxl"""
        start_time = timezone.now()
        
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.utils.dataframe import dataframe_to_rows
            
            # Crear workbook
            wb = Workbook()
            
            # Hoja principal
            ws = wb.active
            ws.title = "Reporte"
            
            # Configurar estilos
            title_font = Font(size=16, bold=True)
            header_font = Font(size=12, bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            
            # Título
            ws['A1'] = f"Reporte: {template.name}"
            ws['A1'].font = title_font
            ws.merge_cells('A1:E1')
            
            # Información del reporte
            row = 3
            ws[f'A{row}'] = "Fecha de generación:"
            ws[f'B{row}'] = timezone.now().strftime('%d/%m/%Y %H:%M')
            row += 1
            
            ws[f'A{row}'] = "Período:"
            ws[f'B{row}'] = f"{data.get('date_from', '')} - {data.get('date_to', '')}"
            row += 1
            
            ws[f'A{row}'] = "Tipo de reporte:"
            ws[f'B{row}'] = template.get_report_type_display()
            row += 2
            
            # Datos principales
            if 'main_data' in data and data['main_data']:
                self._add_excel_data_sheet(wb, data['main_data'], "Datos")
            
            # KPIs
            if 'kpis' in data and data['kpis']:
                self._add_excel_kpis_sheet(wb, data['kpis'])
            
            # Guardar archivo
            wb.save(output_path)
            
            generation_time = (timezone.now() - start_time).total_seconds()
            
            result = self._get_file_info(output_path)
            result.update({
                'mime_type': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                'generation_time': generation_time
            })
            
            return result
            
        except Exception as e:
            logger.error(f"Error generando Excel: {str(e)}")
            raise
    
    def _add_excel_data_sheet(self, wb, data, sheet_name):
        """Agrega hoja de datos al archivo Excel"""
        if not data:
            return
        
        # Crear nueva hoja
        ws = wb.create_sheet(title=sheet_name)
        
        # Convertir a DataFrame para facilitar el manejo
        df = pd.DataFrame(data)
        
        # Agregar headers
        for col_num, column_title in enumerate(df.columns, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = column_title
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        
        # Agregar datos
        for row_num, row_data in enumerate(df.itertuples(index=False), 2):
            for col_num, value in enumerate(row_data, 1):
                ws.cell(row=row_num, column=col_num, value=value)
        
        # Ajustar ancho de columnas
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 30)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    def _add_excel_kpis_sheet(self, wb, kpis):
        """Agrega hoja de KPIs al archivo Excel"""
        ws = wb.create_sheet(title="KPIs")
        
        # Headers
        headers = ['KPI', 'Valor', 'Estado', 'Descripción']
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = header
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        
        # Datos de KPIs
        for row_num, kpi in enumerate(kpis, 2):
            ws.cell(row=row_num, column=1, value=kpi.get('name', ''))
            ws.cell(row=row_num, column=2, value=kpi.get('value', ''))
            ws.cell(row=row_num, column=3, value=kpi.get('status', ''))
            ws.cell(row=row_num, column=4, value=kpi.get('description', ''))
        
        # Ajustar anchos
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 30)
            ws.column_dimensions[column_letter].width = adjusted_width


class CSVReportGenerator(BaseReportGenerator):
    """Generador de reportes en CSV"""
    
    def generate(self, template, data: Dict[str, Any], output_path: str) -> Dict[str, Any]:
        """Genera reporte CSV usando pandas"""
        start_time = timezone.now()
        
        try:
            # Preparar datos principales
            main_data = data.get('main_data', [])
            
            if main_data:
                df = pd.DataFrame(main_data)
                df.to_csv(output_path, index=False, encoding='utf-8-sig')
            else:
                # Crear CSV vacío con headers básicos
                df = pd.DataFrame({'Mensaje': ['No hay datos disponibles para el período seleccionado']})
                df.to_csv(output_path, index=False, encoding='utf-8-sig')
            
            generation_time = (timezone.now() - start_time).total_seconds()
            
            result = self._get_file_info(output_path)
            result.update({
                'mime_type': 'text/csv',
                'generation_time': generation_time
            })
            
            return result
            
        except Exception as e:
            logger.error(f"Error generando CSV: {str(e)}")
            raise


class JSONReportGenerator(BaseReportGenerator):
    """Generador de reportes en JSON"""
    
    def generate(self, template, data: Dict[str, Any], output_path: str) -> Dict[str, Any]:
        """Genera reporte JSON"""
        start_time = timezone.now()
        
        try:
            # Preparar estructura del reporte
            report_data = {
                'template': {
                    'name': template.name,
                    'type': template.report_type,
                    'description': template.description,
                },
                'metadata': {
                    'generated_at': timezone.now().isoformat(),
                    'period': {
                        'from': data.get('date_from', ''),
                        'to': data.get('date_to', ''),
                    }
                },
                'data': data.get('main_data', []),
                'kpis': data.get('kpis', []),
                'summary': data.get('summary', {}),
            }
            
            # Convertir Decimal a float para JSON
            def decimal_to_float(obj):
                if isinstance(obj, Decimal):
                    return float(obj)
                elif isinstance(obj, (date, datetime)):
                    return obj.isoformat()
                raise TypeError
            
            # Guardar JSON
            with open(output_path, 'w', encoding='utf-8') as f:
                json.dump(report_data, f, indent=2, ensure_ascii=False, default=decimal_to_float)
            
            generation_time = (timezone.now() - start_time).total_seconds()
            
            result = self._get_file_info(output_path)
            result.update({
                'mime_type': 'application/json',
                'generation_time': generation_time
            })
            
            return result
            
        except Exception as e:
            logger.error(f"Error generando JSON: {str(e)}")
            raise


class ReportDataService:
    """Servicio para obtener datos para reportes"""
    
    def get_data_for_template(self, template, filters: Dict[str, Any]) -> Dict[str, Any]:
        """
        Obtiene los datos necesarios para generar un reporte
        basado en la plantilla y filtros especificados
        """
        from inventory.models import Product, Transaction, Location
        from alerts.models import Alert
        from forecasting.models import DemandForecast
        
        report_type = template.report_type
        date_from = filters.get('date_from')
        date_to = filters.get('date_to')
        company = template.company
        
        data = {
            'date_from': date_from,
            'date_to': date_to,
            'main_data': [],
            'kpis': [],
            'summary': {},
            'charts': []
        }
        
        try:
            if report_type == 'inventory_summary':
                data.update(self._get_inventory_summary_data(company, date_from, date_to))
            
            elif report_type == 'stock_movement':
                data.update(self._get_stock_movement_data(company, date_from, date_to))
            
            elif report_type == 'abc_analysis':
                data.update(self._get_abc_analysis_data(company, date_from, date_to))
            
            elif report_type == 'alerts_summary':
                data.update(self._get_alerts_summary_data(company, date_from, date_to))
            
            elif report_type == 'forecast_accuracy':
                data.update(self._get_forecast_accuracy_data(company, date_from, date_to))
            
            else:
                # Reporte personalizado o no implementado
                data['main_data'] = [{'mensaje': 'Tipo de reporte no implementado'}]
            
            return data
            
        except Exception as e:
            logger.error(f"Error obteniendo datos para reporte {report_type}: {str(e)}")
            return {
                'main_data': [{'error': str(e)}],
                'kpis': [],
                'summary': {},
                'charts': []
            }
    
    def _get_inventory_summary_data(self, company, date_from, date_to):
        """Obtiene datos para reporte de resumen de inventario"""
        from inventory.models import Product, Transaction
        
        # Productos activos
        products = Product.objects.filter(company=company, is_active=True)
        
        main_data = []
        for product in products:
            stock_actual = product.current_stock or 0
            stock_minimo = product.minimum_stock or 0
            stock_maximo = product.maximum_stock or 0
            
            main_data.append({
                'Código': product.code,
                'Producto': product.name,
                'Categoría': product.category.name if product.category else '',
                'Stock Actual': stock_actual,
                'Stock Mínimo': stock_minimo,
                'Stock Máximo': stock_maximo,
                'Valor Total': float(stock_actual * (product.unit_cost or 0)),
                'Estado': 'Crítico' if stock_actual <= stock_minimo else 'Óptimo'
            })
        
        # KPIs
        total_products = products.count()
        total_value = sum(float(p.current_stock * (p.unit_cost or 0)) for p in products)
        critical_products = products.filter(current_stock__lte=models.F('minimum_stock')).count()
        
        kpis = [
            {'name': 'Total Productos', 'value': total_products, 'status': 'green'},
            {'name': 'Valor Total Inventario', 'value': f"${total_value:,.2f}", 'status': 'green'},
            {'name': 'Productos Críticos', 'value': critical_products, 'status': 'red' if critical_products > 0 else 'green'},
        ]
        
        return {
            'main_data': main_data,
            'kpis': kpis,
            'summary': {
                'total_products': total_products,
                'total_value': total_value,
                'critical_products': critical_products
            }
        }
    
    def _get_stock_movement_data(self, company, date_from, date_to):
        """Obtiene datos para reporte de movimientos de stock"""
        from inventory.models import Transaction
        
        transactions = Transaction.objects.filter(
            product__company=company,
            created_at__date__range=[date_from, date_to]
        ).select_related('product', 'created_by').order_by('-created_at')
        
        main_data = []
        for transaction in transactions:
            main_data.append({
                'Fecha': transaction.created_at.strftime('%d/%m/%Y %H:%M'),
                'Producto': transaction.product.name,
                'Tipo': transaction.get_transaction_type_display(),
                'Cantidad': transaction.quantity,
                'Ubicación': transaction.location.name if transaction.location else '',
                'Referencia': transaction.reference or '',
                'Usuario': transaction.created_by.get_full_name() if transaction.created_by else '',
            })
        
        # KPIs
        total_transactions = transactions.count()
        inbound_count = transactions.filter(transaction_type='in').count()
        outbound_count = transactions.filter(transaction_type='out').count()
        
        kpis = [
            {'name': 'Total Movimientos', 'value': total_transactions, 'status': 'green'},
            {'name': 'Entradas', 'value': inbound_count, 'status': 'green'},
            {'name': 'Salidas', 'value': outbound_count, 'status': 'green'},
        ]
        
        return {
            'main_data': main_data,
            'kpis': kpis,
            'summary': {
                'total_transactions': total_transactions,
                'inbound_count': inbound_count,
                'outbound_count': outbound_count
            }
        }
    
    def _get_abc_analysis_data(self, company, date_from, date_to):
        """Obtiene datos para análisis ABC"""
        from inventory.models import Product
        from django.db.models import Sum, F
        
        # Calcular valor total por producto (esto es simplificado)
        products = Product.objects.filter(
            company=company,
            is_active=True
        ).annotate(
            total_value=F('current_stock') * F('unit_cost')
        ).order_by('-total_value')
        
        main_data = []
        for i, product in enumerate(products, 1):
            if i <= len(products) * 0.2:
                category = 'A'
            elif i <= len(products) * 0.5:
                category = 'B'
            else:
                category = 'C'
            
            main_data.append({
                'Producto': product.name,
                'Valor Total': float(product.total_value or 0),
                'Clasificación ABC': category,
                'Porcentaje Acumulado': round((i / len(products)) * 100, 2)
            })
        
        # KPIs
        category_a = len([p for p in main_data if p['Clasificación ABC'] == 'A'])
        category_b = len([p for p in main_data if p['Clasificación ABC'] == 'B'])
        category_c = len([p for p in main_data if p['Clasificación ABC'] == 'C'])
        
        kpis = [
            {'name': 'Productos Categoría A', 'value': category_a, 'status': 'green'},
            {'name': 'Productos Categoría B', 'value': category_b, 'status': 'yellow'},
            {'name': 'Productos Categoría C', 'value': category_c, 'status': 'red'},
        ]
        
        return {
            'main_data': main_data,
            'kpis': kpis,
            'summary': {
                'category_a': category_a,
                'category_b': category_b,
                'category_c': category_c
            }
        }
    
    def _get_alerts_summary_data(self, company, date_from, date_to):
        """Obtiene datos para resumen de alertas"""
        from alerts.models import Alert
        
        alerts = Alert.objects.filter(
            company=company,
            created_at__date__range=[date_from, date_to]
        ).order_by('-created_at')
        
        main_data = []
        for alert in alerts:
            main_data.append({
                'Fecha': alert.created_at.strftime('%d/%m/%Y %H:%M'),
                'Tipo': alert.get_alert_type_display(),
                'Producto': alert.product.name if alert.product else '',
                'Mensaje': alert.message,
                'Prioridad': alert.get_priority_display(),
                'Estado': alert.get_status_display(),
            })
        
        # KPIs
        total_alerts = alerts.count()
        critical_alerts = alerts.filter(priority='critical').count()
        resolved_alerts = alerts.filter(status='resolved').count()
        
        kpis = [
            {'name': 'Total Alertas', 'value': total_alerts, 'status': 'yellow' if total_alerts > 0 else 'green'},
            {'name': 'Alertas Críticas', 'value': critical_alerts, 'status': 'red' if critical_alerts > 0 else 'green'},
            {'name': 'Alertas Resueltas', 'value': resolved_alerts, 'status': 'green'},
        ]
        
        return {
            'main_data': main_data,
            'kpis': kpis,
            'summary': {
                'total_alerts': total_alerts,
                'critical_alerts': critical_alerts,
                'resolved_alerts': resolved_alerts
            }
        }
    
    def _get_forecast_accuracy_data(self, company, date_from, date_to):
        """Obtiene datos para precisión de pronósticos"""
        from forecasting.models import DemandForecast
        
        forecasts = DemandForecast.objects.filter(
            product__company=company,
            created_at__date__range=[date_from, date_to]
        ).select_related('product')
        
        main_data = []
        for forecast in forecasts:
            accuracy = forecast.accuracy_percentage or 0
            main_data.append({
                'Producto': forecast.product.name,
                'Período': f"{forecast.forecast_period_start} - {forecast.forecast_period_end}",
                'Demanda Pronosticada': forecast.predicted_demand,
                'Demanda Real': forecast.actual_demand or 'N/A',
                'Precisión (%)': accuracy,
                'Estado': 'Bueno' if accuracy >= 80 else 'Regular' if accuracy >= 60 else 'Malo'
            })
        
        # KPIs
        avg_accuracy = forecasts.aggregate(avg=models.Avg('accuracy_percentage'))['avg'] or 0
        good_forecasts = forecasts.filter(accuracy_percentage__gte=80).count()
        total_forecasts = forecasts.count()
        
        kpis = [
            {'name': 'Precisión Promedio', 'value': f"{avg_accuracy:.1f}%", 'status': 'green' if avg_accuracy >= 80 else 'yellow'},
            {'name': 'Pronósticos Buenos', 'value': good_forecasts, 'status': 'green'},
            {'name': 'Total Pronósticos', 'value': total_forecasts, 'status': 'green'},
        ]
        
        return {
            'main_data': main_data,
            'kpis': kpis,
            'summary': {
                'avg_accuracy': avg_accuracy,
                'good_forecasts': good_forecasts,
                'total_forecasts': total_forecasts
            }
        }
